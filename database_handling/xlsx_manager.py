"""
This wonderful module exports user's records to excel file
"""
from datetime import datetime
import openpyxl

from database.db import BotDB


class ExcelManager:
    """
    Works with statistics in excel
    """
    def __init__(self):
        self.book = openpyxl.Workbook()
        self.book.remove(self.book.active)
        self.database = BotDB()

    @staticmethod
    def set_format(record):
        """
        Static method for date formatting
        """
        formatted_record = []
        if record:
            formatted_record.append(record[0][2:])
            formatted_record.append(round(float(record[1]), 2))
            date = datetime.strptime(record[2], "%Y-%m-%d %H:%M:%S")
            formatted_record.append(date.strftime("%d-%m-%Y"))
        return formatted_record

    def get_records(self, user_id):
        """
        Extracts income and expenses from db instance
        """
        incomes = []
        expenses = []
        for i in self.database.get_inc_exp(user_id=user_id, operation=1):
            incomes.append(self.set_format(i))
        for i in self.database.get_inc_exp(user_id=user_id, operation=-1):
            expenses.append(self.set_format(i))
        return incomes, expenses

    def fill_excel(self, incomes, expenses):
        """
        Fills the Excel sheet with exciting exctracted information
        """
        for row in incomes:
            self.book["Доходы"].append(row)
        for row in expenses:
            self.book["Расходы"].append(row)

    def save_excel(self, user_id):
        """
        Creates and saves the Excel sheet under the name FinanceHunterBot.xlsx
        """
        self.book.create_sheet("Доходы")
        self.book.create_sheet("Расходы")

        columns = ['A', 'B', 'C']
        for sheet in self.book.worksheets:
            for col in columns:
                sheet.column_dimensions[col].width = 20
            sheet["A1"].value = "Категория"
            sheet["B1"].value = f"Сумма ({self.database.get_currency(user_id=user_id)})"
            sheet["C1"].value = "Дата"

        records = self.get_records(user_id=user_id)
        self.fill_excel(records[0], records[1])

        return self.book.save('FinanceHunterBot.xlsx')
